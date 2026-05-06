import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
import os, re, io, zipfile, hashlib
from datetime import datetime, date

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE PÁGINA
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Generador Certificados SLB",
    page_icon="📄",
    layout="centered",
)

# ─────────────────────────────────────────────────────────────────────────────
# USUARIOS AUTORIZADOS  (usuario: hash de contraseña)
# ─────────────────────────────────────────────────────────────────────────────
def _hash(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

USUARIOS = {
    "AMASTRACCI": _hash("sullair2026"),
    "JICUZA":     _hash("sullair2026"),
    "ETOLOSA":    _hash("sullair2026"),
    "FCENDRA":    _hash("sullair2026"),
}

# ─────────────────────────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────────────────────────
def mostrar_login():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Sans:wght@300;400;500;600&display=swap');
    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
    .login-box {
        max-width: 380px;
        margin: 60px auto 0;
        background: #fff;
        border-radius: 16px;
        padding: 40px 36px;
        box-shadow: 0 4px 24px rgba(0,0,0,.10);
        border: 1.5px solid #E0E0E0;
        text-align: center;
    }
    .login-titulo {
        font-family: 'Syne', sans-serif;
        font-size: 1.4rem;
        font-weight: 800;
        color: #00783C;
        margin-bottom: 4px;
    }
    .login-sub { font-size: .82rem; color: #888; margin-bottom: 28px; }
    </style>
    """, unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        st.markdown('<div class="login-box">', unsafe_allow_html=True)
        st.markdown('<p class="login-titulo">📄 Sullair Argentina</p>', unsafe_allow_html=True)
        st.markdown('<p class="login-sub">Generador de Certificados SLB</p>', unsafe_allow_html=True)

        usuario = st.text_input("Usuario", placeholder="Ej: JICUZA").strip().upper()
        password = st.text_input("Contraseña", type="password", placeholder="••••••••••")

        if st.button("Ingresar", use_container_width=True):
            if usuario in USUARIOS and USUARIOS[usuario] == _hash(password):
                st.session_state['logged_in'] = True
                st.session_state['usuario'] = usuario
                st.rerun()
            else:
                st.error("Usuario o contraseña incorrectos.")

        st.markdown('</div>', unsafe_allow_html=True)

# Verificar sesión
if 'logged_in' not in st.session_state or not st.session_state['logged_in']:
    mostrar_login()
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# CSS PERSONALIZADO
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Sans:wght@300;400;500;600&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

.titulo-app {
    font-family: 'Syne', sans-serif;
    font-size: 1.6rem;
    font-weight: 800;
    color: #00783C;
    margin-bottom: 0;
}
.subtitulo-app {
    font-size: .88rem;
    color: #888;
    margin-bottom: 24px;
}
.seccion {
    background: #fff;
    border: 1.5px solid #E0E0E0;
    border-radius: 14px;
    padding: 22px 26px;
    margin-bottom: 18px;
    box-shadow: 0 2px 10px rgba(0,0,0,.05);
}
.seccion h3 {
    font-family: 'Syne', sans-serif;
    color: #00783C;
    font-size: 1rem;
    margin-bottom: 14px;
}
.stat-box {
    background: #E8F5EE;
    border: 1.5px solid #b2dfc5;
    border-radius: 10px;
    padding: 14px;
    text-align: center;
}
.stat-num {
    font-family: 'Syne', sans-serif;
    font-size: 2rem;
    font-weight: 800;
    color: #00783C;
    display: block;
}
.stat-lbl { font-size: .75rem; color: #555; }
.stButton > button {
    background: #00783C !important;
    color: white !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 700 !important;
    border-radius: 10px !important;
    border: none !important;
    padding: 12px 28px !important;
    width: 100%;
}
.stButton > button:hover { background: #005a2c !important; }
div[data-testid="stDownloadButton"] > button {
    background: white !important;
    color: #00783C !important;
    border: 2px solid #00783C !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 700 !important;
    border-radius: 10px !important;
    width: 100%;
}
div[data-testid="stDownloadButton"] > button:hover {
    background: #00783C !important;
    color: white !important;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# RUTAS DE IMÁGENES
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(BASE_DIR, "assets", "logo_sullair.jpg")
FIRMA_PATH= os.path.join(BASE_DIR, "assets", "firma_mastracci.jpg")

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
VERDE = colors.Color(113/255, 175/255, 71/255)
NEGRO = colors.black

def limpiar(s):
    return re.sub(r'[\\/*?:"<>|]', '', str(s).strip()).replace('  ', ' ')

def fmt_fecha(dt):
    if dt is None: return ''
    if isinstance(dt, (datetime, date)):
        return dt.strftime('%d/%m/%Y')
    return str(dt)

def extraer_modelo_interno(campo_m):
    partes = [p.strip().lstrip('#') for p in str(campo_m).split('/')]
    interno = ''
    for p in partes:
        if re.match(r'^(E0\d{5}|A0\d{5})', p, re.I):
            interno = p.strip(); break
    if not interno:
        for p in partes:
            if re.match(r'^[34]\d{3,}$', p.strip()):
                interno = p.strip(); break
    modelo = partes[0].strip() if partes else ''
    return modelo, interno

def calcular_cantidad(desde, hasta):
    if desde is None or hasta is None: return ''
    if isinstance(desde, datetime): desde = desde.date()
    if isinstance(hasta, datetime): hasta = hasta.date()
    dias = (hasta - desde).days + 1
    return '1 (MES)' if dias >= 28 else f'{dias} (días)'

# ─────────────────────────────────────────────────────────────────────────────
# GENERADOR PDF — idéntico al script que ya funciona
# ─────────────────────────────────────────────────────────────────────────────
def generar_pdf(cert_num, segmento, cvu, periodo, equipo,
                modelo, interno, desde, hasta, cantidad,
                cliente, sucursal, direccion):

    PAGE_W, PAGE_H = A4
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)

    BOX_L = 1.75 * cm
    BOX_R = PAGE_W - 1.75 * cm
    BOX_T = PAGE_H - 2.0  * cm
    BOX_B = PAGE_H - 15.8 * cm
    BOX_W = BOX_R - BOX_L

    # Recuadro exterior
    c.setStrokeColor(NEGRO)
    c.setLineWidth(1.2)
    c.rect(BOX_L, BOX_B, BOX_W, BOX_T - BOX_B, stroke=1, fill=0)

    # Logo
    LOGO_W = 4.6 * cm
    LOGO_H = 1.35 * cm
    if os.path.exists(LOGO_PATH):
        c.drawImage(LOGO_PATH,
                    BOX_R - LOGO_W - 0.35*cm,
                    BOX_T - LOGO_H - 0.55*cm,
                    width=LOGO_W, height=LOGO_H,
                    preserveAspectRatio=True, mask='auto')

    # Título
    TX = BOX_L + 0.65*cm
    c.setFont('Helvetica-Bold', 11.5)
    c.setFillColor(NEGRO)
    c.drawString(TX, BOX_T - 1.15*cm, 'CERTIFICACIÓN DE SERVICIOS')
    c.setFont('Helvetica', 9.5)
    c.drawString(TX, BOX_T - 1.75*cm, 'Alquiler de equipos')

    # Línea separadora
    SEP_Y = BOX_T - 2.2*cm
    c.setLineWidth(0.5)
    c.line(BOX_L, SEP_Y, BOX_R, SEP_Y)

    # Encabezado
    LBL_X = TX
    VAL_X = LBL_X + 3.2*cm
    MID_X = BOX_L + BOX_W * 0.50
    ROW_H = 0.58*cm
    Y0    = SEP_Y - 0.68*cm

    def lbl(y, txt):
        c.setFont('Helvetica-Bold', 9.5)
        c.setFillColor(NEGRO)
        c.drawString(LBL_X, y, txt)

    def val(y, txt, x=None):
        c.setFont('Helvetica', 9.5)
        c.setFillColor(NEGRO)
        c.drawString(x if x is not None else VAL_X, y, str(txt))

    lbl(Y0,            'Certificado N°:');  val(Y0,            str(cert_num))
    val(Y0,            sucursal,            x=MID_X)

    lbl(Y0 - ROW_H,    'Cliente:');         val(Y0 - ROW_H,    cliente)
    val(Y0 - ROW_H,    direccion,           x=MID_X)

    lbl(Y0 - 2*ROW_H,  'Periodo:');         val(Y0 - 2*ROW_H,  periodo)

    lbl(Y0 - 3*ROW_H,  'CVU:');             val(Y0 - 3*ROW_H,  cvu.strip())
    c.setFont('Helvetica-Bold', 9.5)
    c.setFillColor(NEGRO)
    c.drawString(MID_X, Y0 - 3*ROW_H, 'Fecha Certif.:')
    c.setFont('Helvetica', 9.5)
    c.drawString(MID_X + 3.05*cm, Y0 - 3*ROW_H,
                 datetime.today().strftime('%d/%m/%Y'))

    # Tabla
    TAB_TOP = SEP_Y - 3.45*cm
    TAB_L   = BOX_L + 0.35*cm
    TAB_W   = BOX_W - 0.7*cm

    CW = [TAB_W*0.358, TAB_W*0.118, TAB_W*0.148,
          TAB_W*0.133, TAB_W*0.133, TAB_W*0.110]

    data = [
        ['Equipo', 'Modelo', 'Interno', 'Desde', 'Hasta', 'Cant.'],
        [equipo, modelo, interno,
         fmt_fecha(desde), fmt_fecha(hasta), cantidad],
        ['', '', '', '', '', ''],
    ]

    t = Table(data, colWidths=CW, rowHeights=[16, 15, 9])
    t.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), VERDE),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 8.5),
        ('ALIGN',       (0,0), (-1,0), 'CENTER'),
        ('VALIGN',      (0,0), (-1,0), 'MIDDLE'),
        ('FONTNAME',    (0,1), (-1,1), 'Helvetica'),
        ('FONTSIZE',    (0,1), (-1,1), 8.5),
        ('ALIGN',       (1,1), (-1,1), 'CENTER'),
        ('ALIGN',       (0,1), (0,1),  'LEFT'),
        ('LEFTPADDING', (0,1), (0,1),  4),
        ('VALIGN',      (0,1), (-1,1), 'MIDDLE'),
        ('FONTSIZE',    (0,2), (-1,2), 7),
        ('BOX',         (0,0), (-1,-1), 0.8, NEGRO),
        ('INNERGRID',   (0,0), (-1,-1), 0.5, NEGRO),
    ]))
    t.wrapOn(c, TAB_W, 200)
    t.drawOn(c, TAB_L, TAB_TOP - t._height)

    # Firma
    FIRMA_W = 4.8 * cm
    FIRMA_H = 2.2 * cm
    FIRMA_X = BOX_R - FIRMA_W - 0.4*cm
    FIRMA_Y = BOX_B + 0.7*cm
    if os.path.exists(FIRMA_PATH):
        c.drawImage(FIRMA_PATH, FIRMA_X, FIRMA_Y,
                    width=FIRMA_W, height=FIRMA_H,
                    preserveAspectRatio=True, mask='auto')

    # Texto exterior
    c.setFont('Helvetica', 8.5)
    c.setFillColor(NEGRO)
    c.drawRightString(BOX_R, BOX_B - 0.55*cm, 'SULLAIR ARGENTINA S.A')

    c.save()
    buf.seek(0)
    return buf.read()

# ─────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────
def procesar_planilla(archivo_bytes, cert_inicio, periodo_manual,
                      cliente, sucursal, direccion, progress_bar, status_text):

    # Leer planilla con openpyxl (data_only para valores)
    wb_data = load_workbook(io.BytesIO(archivo_bytes), data_only=True)
    ws_data = wb_data['CONSOLIDADO']

    # Planilla de salida — copia exacta sin tocar nada
    wb_out  = load_workbook(io.BytesIO(archivo_bytes))
    ws_out  = wb_out['CONSOLIDADO']

    # Período
    p_mes  = ws_data['B3'].value or ''
    p_anio = ws_data['B4'].value or ''
    periodo_detectado = f"{str(p_mes).strip()} {str(p_anio).strip()}".strip()
    periodo = periodo_manual.strip() if periodo_manual.strip() else periodo_detectado

    cert_num  = cert_inicio
    generados = 0
    errores   = []
    log_lines = []

    # Contar filas primero para la barra de progreso
    filas_alq = []
    for row_num in range(9, 200):
        tipo  = ws_data[f'H{row_num}'].value
        seg   = ws_data[f'G{row_num}'].value
        equip = ws_data[f'L{row_num}'].value
        if tipo is None and seg is None and equip is None:
            break
        if str(tipo).strip().upper() == 'ALQUILER':
            filas_alq.append(row_num)

    # ZIP en memoria
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for idx, row_num in enumerate(filas_alq):
            seg   = ws_data[f'G{row_num}'].value
            cvu   = ws_data[f'D{row_num}'].value
            equip = ws_data[f'L{row_num}'].value
            m_col = ws_data[f'M{row_num}'].value
            desde = ws_data[f'S{row_num}'].value
            hasta = ws_data[f'T{row_num}'].value

            modelo, interno = extraer_modelo_interno(m_col or '')
            if not interno:
                msg = f"Fila {row_num}: sin interno en '{m_col}'"
                errores.append(msg)
                log_lines.append(f"⚠️  {msg}")
                continue

            cvu_val   = str(cvu).strip()   if cvu   else 'S/N'
            seg_val   = str(seg).strip()   if seg   else 'SIN_SEG'
            equip_val = str(equip).strip() if equip else interno
            cantidad  = calcular_cantidad(desde, hasta)

            pdf_bytes = generar_pdf(
                cert_num, seg_val, cvu_val, periodo, equip_val,
                modelo, interno, desde, hasta, cantidad,
                cliente, sucursal, direccion
            )

            nombre = f"{limpiar(seg_val)} - {limpiar(interno)} - {cert_num}.pdf"
            zf.writestr(nombre, pdf_bytes)

            # Escribir N° en planilla de salida
            ws_out[f'O{row_num}'] = cert_num

            log_lines.append(f"✅ Cert {cert_num:04d} | {seg_val:<20} | {interno}")
            cert_num  += 1
            generados += 1

            # Actualizar progreso
            pct = int((idx + 1) / len(filas_alq) * 100)
            progress_bar.progress(pct)
            status_text.text(f"Generando {idx+1} de {len(filas_alq)}...")

    zip_buf.seek(0)
    zip_bytes = zip_buf.read()

    # Guardar Excel de salida en memoria
    xlsx_buf = io.BytesIO()
    wb_out.save(xlsx_buf)
    xlsx_buf.seek(0)
    xlsx_bytes = xlsx_buf.read()

    return {
        'generados':  generados,
        'errores':    errores,
        'log':        log_lines,
        'zip_bytes':  zip_bytes,
        'xlsx_bytes': xlsx_bytes,
        'periodo':    periodo,
        'cert_inicio': cert_inicio,
        'cert_fin':   cert_num - 1,
    }

# ─────────────────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────────────────
col_title, col_logout = st.columns([4, 1])
with col_title:
    st.markdown('<p class="titulo-app">📄 Generador de Certificados SLB</p>', unsafe_allow_html=True)
    st.markdown(f'<p class="subtitulo-app">Sullair Argentina S.A. — Vaca Muerta &nbsp;|&nbsp; 👤 {st.session_state.get("usuario","")}</p>', unsafe_allow_html=True)
with col_logout:
    st.write("")
    if st.button("🚪 Salir", use_container_width=True):
        st.session_state['logged_in'] = False
        st.session_state['usuario'] = ''
        st.rerun()

# ── Sección 1: Configuración ──────────────────────────────────────────────────
st.markdown('<div class="seccion"><h3>⚙️ Configuración del período</h3>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    cert_inicio = st.number_input("N° de certificado inicial", min_value=1, value=1000, step=1)
with col2:
    periodo_manual = st.text_input("Período", placeholder="Se detecta del Excel automáticamente")

col3, col4 = st.columns(2)
with col3:
    cliente  = st.text_input("Cliente", value="Schlumberger")
with col4:
    sucursal = st.text_input("Sucursal", value="Sucursal NQN y Valle")

direccion = st.text_input("Dirección cliente", value="Industriales Neuquinos 2920 - PIN OESTE")
st.markdown('</div>', unsafe_allow_html=True)

# ── Sección 2: Cargar Excel ───────────────────────────────────────────────────
st.markdown('<div class="seccion"><h3>📊 Cargar Planilla de Certificación</h3>', unsafe_allow_html=True)
archivo = st.file_uploader(
    "Arrastrá o seleccioná el archivo Excel",
    type=["xlsx", "xls"],
    label_visibility="collapsed"
)
if archivo:
    st.success(f"✅ **{archivo.name}** cargado correctamente")
st.markdown('</div>', unsafe_allow_html=True)

# ── Sección 3: Generar ───────────────────────────────────────────────────────
st.markdown('<div class="seccion"><h3>⚡ Generar certificados</h3>', unsafe_allow_html=True)

if st.button("⚡ Generar PDFs + actualizar planilla", disabled=(archivo is None)):
    progress_bar = st.progress(0)
    status_text  = st.empty()

    with st.spinner("Procesando..."):
        resultado = procesar_planilla(
            archivo_bytes  = archivo.read(),
            cert_inicio    = int(cert_inicio),
            periodo_manual = periodo_manual,
            cliente        = cliente,
            sucursal       = sucursal,
            direccion      = direccion,
            progress_bar   = progress_bar,
            status_text    = status_text,
        )

    progress_bar.progress(100)
    status_text.text("✅ Completado")

    # Guardar en session_state para los botones de descarga
    st.session_state['resultado'] = resultado

    # Log
    with st.expander("📋 Detalle del proceso", expanded=False):
        for line in resultado['log']:
            st.text(line)
        if resultado['errores']:
            st.warning("Advertencias:\n" + "\n".join(resultado['errores']))

st.markdown('</div>', unsafe_allow_html=True)

# ── Sección 4: Descargar ─────────────────────────────────────────────────────
if 'resultado' in st.session_state:
    r = st.session_state['resultado']
    st.markdown('<div class="seccion"><h3>📥 Descargar resultados</h3>', unsafe_allow_html=True)

    # Stats
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="stat-box"><span class="stat-num">{r["generados"]}</span><span class="stat-lbl">PDFs generados</span></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="stat-box"><span class="stat-num">{r["cert_inicio"]}</span><span class="stat-lbl">N° inicial</span></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="stat-box"><span class="stat-num">{r["cert_fin"]}</span><span class="stat-lbl">N° final</span></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="stat-box"><span class="stat-num" style="font-size:1.1rem">{r["periodo"]}</span><span class="stat-lbl">Período</span></div>', unsafe_allow_html=True)

    st.write("")

    col_a, col_b = st.columns(2)
    with col_a:
        st.download_button(
            label="📦 Descargar ZIP (PDFs)",
            data=r['zip_bytes'],
            file_name=f"Certificados_SLB_{r['cert_inicio']}_{r['cert_fin']}.zip",
            mime="application/zip",
        )
    with col_b:
        st.download_button(
            label="📋 Descargar Planilla completada",
            data=r['xlsx_bytes'],
            file_name="Planilla_Certificacion_SLB_COMPLETADA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.markdown('</div>', unsafe_allow_html=True)

# Footer
st.markdown("---")
st.markdown('<p style="text-align:center;color:#aaa;font-size:.75rem">Sullair Argentina S.A. · Generador automático de certificados SLB · v3.0</p>', unsafe_allow_html=True)
